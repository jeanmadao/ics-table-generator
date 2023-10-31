import os
import requests
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from dotenv import load_dotenv
from openpyxl import Workbook

from style import style_worksheet
from functions import parse_timedate

load_dotenv()
wb = Workbook()
ws = wb.active

ws.append(["Date", "Location", "Start", "End", "Break", "Hours"])

ICS = requests.get(os.getenv('SCHEDULE_ICS_URL')).text.split('\r\n')

lines_nb = len(ICS)
workdays = []
i = 0
while i < lines_nb:
    if ICS[i].startswith('DTSTART;'):
        start_date, start_time = parse_timedate(ICS[i])
        if start_date.month == 10:
            while not ICS[i].startswith('DTEND;'):
                i += 1
            end_date, end_time = parse_timedate(ICS[i])
            workdays.append((start_date, "Louise", start_time, end_time))
    i += 1

workdays.sort(key=lambda workday: workday[0])

row = 2
for workday in workdays:
    start_date, location, start_time, end_time = workday
    ws.append([start_date.strftime('%d/%m/%Y'), location, start_time, end_time, f'=IF((D{row} - C{row}) < TIME(5, 0, 0), "00:00:00", "00:30:00"', f'=D{row} - C{row} - E{row}'])
    row += 1

ws[f'A{row}'] = "Total"
ws[f'F{row}'] = f'=SUM(F2:F{row - 1})'

style_worksheet(ws, row, 5)

wb.save(os.getenv('OUTPUT_PATH'))
