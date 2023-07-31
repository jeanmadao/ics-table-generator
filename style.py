from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from style_variables import *

def style_worksheet(ws, rows, columns):
    ws.row_dimensions[1].font = Font(bold=True,
                                     color=WHITE,
                                     name=FONT)
    ws.row_dimensions[1].fill = PatternFill(patternType='solid',
                                            fgColor=RED)
    ws.column_dimensions['A'].width = DATE_CELL_WIDTH

    ws.row_dimensions[rows].font = Font(bold=True,
                                        color=WHITE,
                                        name=FONT)
    ws.row_dimensions[rows].fill = PatternFill(patternType='solid',
                                               fgColor=BLUE)
